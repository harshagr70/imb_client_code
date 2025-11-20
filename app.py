import streamlit as st
import pandas as pd
import asyncio
import tempfile
import os
from pathlib import Path
import json
import re
from io import BytesIO
from datetime import datetime
from collections import defaultdict
from typing import Any, Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import sys
import importlib.util

# Import pipeline components
from helpers import llm_validator_new
from helpers.attention import *
from helpers.llm_validator_new import *
from helpers.test_json_operator import *
from llama_index.core import SimpleDirectoryReader
from llama_parse import LlamaParse

# Page configuration
st.set_page_config(
    page_title="Financial Statement Extractor",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for better styling
st.markdown("""
    <style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
    }
    .statement-header {
        font-size: 1.8rem;
        font-weight: bold;
        color: #2c3e50;
        margin-top: 2rem;
        margin-bottom: 1rem;
        padding-bottom: 0.5rem;
        border-bottom: 2px solid #3498db;
    }
    .info-box {
        background-color: #e8f4f8;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 1rem 0;
    }
    .hero-card {
        background: linear-gradient(120deg, #1f77b4 0%, #5ad1f9 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 1rem;
        margin-bottom: 1.5rem;
        box-shadow: 0 10px 30px rgba(0,0,0,0.15);
    }
    .metric-card {
        background-color: #ffffff;
        border-radius: 0.75rem;
        padding: 1rem;
        border: 1px solid #eef2f7;
        box-shadow: 0 4px 15px rgba(0,0,0,0.05);
    }
    .metric-value {
        font-size: 1.6rem;
        font-weight: bold;
        color: #1f77b4;
    }
    .metric-label {
        font-size: 0.95rem;
        color: #5f6b7c;
    }
    </style>
""", unsafe_allow_html=True)

# Initialize session state
if 'processing_complete' not in st.session_state:
    st.session_state.processing_complete = False
if 'extracted_data' not in st.session_state:
    st.session_state.extracted_data = {}
if 'uploaded_file' not in st.session_state:
    st.session_state.uploaded_file = None
if 'latest_logs' not in st.session_state:
    st.session_state.latest_logs = []
if 'latest_progress' not in st.session_state:
    st.session_state.latest_progress = 0

def normalize_year_key(key: Any) -> Optional[str]:
    """Normalize a year/date string into a 4-digit year."""
    if key is None:
        return None

    if isinstance(key, int):
        if 1800 <= key <= 2200:
            return str(key)
        return None

    if isinstance(key, float):
        if key.is_integer() and 1800 <= int(key) <= 2200:
            return str(int(key))
        return None

    candidate = str(key).strip()
    if not candidate:
        return None

    if candidate.isdigit() and len(candidate) == 4:
        return candidate

    # Try to parse common date formats
    for fmt in ("%b %d, %Y", "%B %d, %Y", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(candidate, fmt)
            return str(dt.year)
        except ValueError:
            continue

    match = re.search(r"(18|19|20|21)\d{2}", candidate)
    if match:
        return match.group(0)

    return None

def build_period_columns(periods: List[Dict[str, Any]], rows: List[Dict[str, Any]]) -> List[str]:
    """Return ordered, unique normalized year keys present across periods and row values."""
    ordered_years: List[str] = []
    seen = set()

    def add_year(raw_value: Any):
        norm = normalize_year_key(raw_value)
        if norm and norm not in seen:
            ordered_years.append(norm)
            seen.add(norm)

    for period in periods or []:
        add_year(period.get("year"))
        add_year(period.get("date"))
        add_year(period.get("label"))

    for row in rows or []:
        values = row.get("values") or {}
        for key in values.keys():
            add_year(key)

    return ordered_years

def extract_metric_value(statement: Dict[str, Any], keywords: List[str]) -> Optional[float]:
    """Find the first matching numeric value for provided keywords."""
    if not statement:
        return None
    rows = statement.get("rows") or []
    period_keys = build_period_columns(statement.get("periods"), rows)
    if not period_keys:
        return None

    latest_period = period_keys[-1]
    for row in rows:
        label = (row.get("label") or "").lower()
        if any(keyword in label for keyword in keywords):
            value = row.get("values", {}).get(latest_period)
            if isinstance(value, (int, float)):
                return float(value)
    return None

def format_currency(value: Optional[float]) -> str:
    if value is None:
        return "—"
    abs_val = abs(value)
    if abs_val >= 1_000_000_000:
        return f"{value/1_000_000_000:.1f}B"
    if abs_val >= 1_000_000:
        return f"{value/1_000_000:.1f}M"
    if abs_val >= 1_000:
        return f"{value/1_000:.1f}K"
    return f"{value:,.0f}"

def render_statement_chart(statement: Dict[str, Any], title: str):
    rows = statement.get("rows") or []
    period_keys = build_period_columns(statement.get("periods"), rows)
    if len(period_keys) < 2:
        return

    revenue_row = None
    for row in rows:
        label = (row.get("label") or "").lower()
        if any(term in label for term in ["revenue", "sales", "net sales"]):
            revenue_row = row
            break

    if not revenue_row:
        return

    values = []
    for period in period_keys:
        val = revenue_row.get("values", {}).get(period)
        values.append(val if isinstance(val, (int, float)) else None)

    if all(v is None for v in values):
        return

    chart_df = pd.DataFrame({"period": period_keys, "value": values})
    chart_df = chart_df.dropna()
    if len(chart_df) < 2:
        return

    st.line_chart(chart_df.set_index("period"), height=220, use_container_width=True)

def process_pdf_pipeline(
    pdf_path: str,
    llama_api_key: str,
    openai_api_key: str,
    log_callback=None,
    progress_callback=None,
):
    """
    Main pipeline function that processes PDF and extracts financial statements.
    This function encapsulates all the pipeline logic.
    """
    total_steps = 6
    current_step = 0

    def log(message: str, status: str = "info"):
        if log_callback:
            log_callback(message, status)

    def advance():
        nonlocal current_step
        current_step += 1
        if progress_callback:
            progress_callback(current_step, total_steps)

    if progress_callback:
        progress_callback(0, total_steps)

    try:
        # Step 1: Parse PDF using parser API
        log("Parsing PDF with parser API...", "running")
        with st.spinner("📄 Parsing PDF document..."):
            parser = LlamaParse(
                api_key=llama_api_key,
                premium_mode=True,
                adaptive_long_table=True,
                result_type="markdown",
            )
            file_extractor = {".pdf": parser}
            documents = SimpleDirectoryReader(
                input_files=[pdf_path],
                file_extractor=file_extractor
            ).load_data()
        log("PDF parsing complete.", "success")
        advance()

        # Step 2: Prefilter pages (use enhanced prefilter if available)
        log("Filtering relevant pages...", "running")
        with st.spinner("🔍 Filtering relevant pages..."):
            false_pass_indices = []
            try:
                prefilter_path = Path(__file__).parent / "helpers" / "prefilter_rmd_statement.py"
                if prefilter_path.exists():
                    spec = importlib.util.spec_from_file_location("prefilter_rmd_statement", prefilter_path)
                    pr = importlib.util.module_from_spec(spec)
                    spec.loader.exec_module(pr)
                    filter_func = getattr(pr, "prefilter_statement_page_from_rmd", None)
                    if callable(filter_func):
                        for idx, page in enumerate(documents):
                            result = filter_func(page.text)
                            if result.get("pass", False):
                                false_pass_indices.append(idx)
                    else:
                        false_pass_indices = list(range(len(documents)))
                else:
                    # If prefilter module doesn't exist, include all pages
                    false_pass_indices = list(range(len(documents)))
            except (ImportError, AttributeError, Exception):
                # If prefilter module doesn't exist or fails, include all pages
                false_pass_indices = list(range(len(documents)))
                log("Prefilter unavailable. Including all pages for evaluation.", "warning")

            # Transform documents
            transformed_documents = transform_selected_documents(documents, false_pass_indices)
        log("Relevant pages filtered.", "success")
        advance()

        # Step 3: Process pages for attention
        log("Detecting financial statement pages...", "running")
        with st.spinner("🎯 Identifying financial statement pages..."):
            results = process_pages_for_attention(transformed_documents, is_not=False)
            included_pages = get_included_pages(results)
        log("Statement pages identified.", "success")
        advance()

        # Step 4: Sort and order pages
        log("Organizing statements and table order...", "running")
        with st.spinner("📋 Organizing financial statements..."):
            order = {'income statement': 0, 'balance sheet': 1, 'cashflow': 2}
            sorted_pages = dict(sorted(included_pages.items(), key=lambda x: order.get(x[1], 99)))
            page_nums = list(sorted_pages.keys())
            as_dicts = get_ordered_dicts_from_pages(transformed_documents, page_nums)
        log("Statements organized.", "success")
        advance()

        # Step 5: Validate and extract structured data
        log("Validating and structuring tables with LLM...", "running")
        with st.spinner("✅ Extracting and validating financial data..."):
            vt = LLMOnlyFinancialTableValidator(
                OpenAILLM(api_key=openai_api_key),
                model="gpt-4o-mini"  # Using gpt-4o-mini as gpt-4.1-mini may not be available
            )

            # Convert to async
            dict_results = asyncio.run(
                run_validator_on_pages_llm(vt, as_dicts, max_concurrency=3)
            )
        log("Tables validated successfully.", "success")
        advance()

        # Step 6: Organize results by statement type
        log("Finalizing structured output...", "running")
        organized_data = {
            'income_statement': None,
            'balance_sheet': None,
            'cash_flow': None
        }

        for result in dict_results:
            if result.data and not result.error:
                statement_type = result.data.get('statement_type', '').lower()
                if statement_type == 'income_statement':
                    organized_data['income_statement'] = result.data
                elif statement_type == 'balance_sheet':
                    organized_data['balance_sheet'] = result.data
                elif statement_type == 'cash_flow':
                    organized_data['cash_flow'] = result.data

        log("Financial statement data ready.", "success")
        advance()
        if progress_callback:
            progress_callback(total_steps, total_steps)

        return organized_data, None

    except Exception as e:
        log(f"Pipeline error: {e}", "error")
        if progress_callback:
            progress_callback(total_steps, total_steps)
        return None, str(e)

def json_to_dataframe(data: Dict[str, Any]) -> Optional[pd.DataFrame]:
    """Convert JSON financial data to pandas DataFrame for display with normalized columns."""
    if not data or not isinstance(data, dict):
        return None

    rows = data.get("rows") or []
    periods = data.get("periods") or []

    if not rows:
        return None

    period_keys = build_period_columns(periods, rows)
    if not period_keys:
        return None

    columns = ["Line Item"] + period_keys
    table_data: List[List[str]] = []

    for row in rows:
        if not isinstance(row, dict):
            continue

        label = row.get("label", "")
        is_section = row.get("is_section", False)
        display_label = f"**{label}**" if is_section else label
        values = row.get("values") or {}

        normalized_values: Dict[str, Any] = {}
        for key, value in values.items():
            norm = normalize_year_key(key) or str(key).strip()
            if norm and norm not in normalized_values:
                normalized_values[norm] = value

        row_data = [display_label]
        for key in period_keys:
            value = normalized_values.get(key)
            if value is None:
                row_data.append("-")
                continue

            try:
                num_value = float(value)
                formatted = f"{int(num_value):,}" if num_value == int(num_value) else f"{num_value:,.2f}"
            except (ValueError, TypeError):
                formatted = str(value)
            row_data.append(formatted)

        table_data.append(row_data)

    if not table_data:
        return None

    return pd.DataFrame(table_data, columns=columns)

def merge_statements(statements: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Group statements by type and merge rows/periods similar to reference exporter."""
    grouped: Dict[str, Dict[str, Any]] = defaultdict(
        lambda: {"rows": [], "periods": [], "units": None, "statement_type": None}
    )

    for stmt in statements:
        stype = stmt.get("statement_type") or "unknown"
        bucket = grouped[stype]
        bucket["statement_type"] = stype
        bucket["rows"].extend(stmt.get("rows") or [])

        existing_years = {p.get("year") for p in bucket["periods"]}
        for period in stmt.get("periods") or []:
            year = period.get("year")
            if year not in existing_years:
                bucket["periods"].append(period)
                existing_years.add(year)

        if bucket["units"] is None:
            bucket["units"] = stmt.get("units")

    return list(grouped.values())

def statements_to_excel(statements: List[Dict[str, Any]]) -> BytesIO:
    """Export statements to Excel following the reference layout."""
    wb = Workbook()
    ws = wb.active
    ws.title = "All Financial Tables"
    ws.freeze_panes = "A5"

    title_font = Font(name="Arial", size=14, bold=True)
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    navy_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
    bold_font = Font(name="Arial", bold=True)
    number_font = Font(name="Arial", color="000080")
    center_align = Alignment(horizontal="center", vertical="center")

    def is_validation_schema(item: Dict[str, Any]) -> bool:
        return (
            isinstance(item, dict)
            and "statement_type" in item
            and "periods" in item
            and "rows" in item
        )

    def statement_title(stype: str) -> str:
        mapping = {
            "income_statement": "INCOME STATEMENT",
            "balance_sheet": "BALANCE SHEET",
            "cash_flow": "CASH FLOW STATEMENT",
            "cash_flow_statement": "CASH FLOW STATEMENT",
        }
        return mapping.get(stype, (stype or "UNTITLED")).replace("_", " ").upper()

    normalized_years = set()
    merged_statements: List[Dict[str, Any]] = []

    if statements and is_validation_schema(statements[0]):
        merged_statements = merge_statements(statements)
        scan_source = merged_statements
    else:
        scan_source = statements

    for entry in scan_source:
        for period in entry.get("periods", []):
            normalized_years.add(normalize_year_key(period.get("year")))
        for row in entry.get("rows", []):
            for key in (row.get("values") or {}).keys():
                normalized_years.add(normalize_year_key(key))

    normalized_years = {year for year in normalized_years if year}
    sorted_years = sorted(normalized_years, key=int) if normalized_years else []

    num_cols = 1 + len(sorted_years)
    if num_cols < 2:
        num_cols = 2

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    title_cell = ws.cell(row=2, column=1, value="HISTORICAL FINANCIAL STATEMENTS")
    title_cell.font = title_font
    title_cell.alignment = center_align

    ws.cell(row=3, column=1, value="Fiscal Year Ended").font = Font(name="Arial", size=11)

    header_row = ["Label"] + sorted_years if sorted_years else ["Label", "Value"]
    for col_idx, col_val in enumerate(header_row, 1):
        cell = ws.cell(row=4, column=col_idx, value=col_val)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = center_align

    current_row = 5

    def write_section_row(label: str):
        nonlocal current_row
        cell = ws.cell(row=current_row, column=1, value=label)
        cell.font = bold_font
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        current_row += 1

    target_statements = merged_statements if merged_statements else statements

    if not target_statements:
        ws.cell(row=current_row, column=1, value="No financial statements available.")
    else:
        for stmt in target_statements:
            title = statement_title(stmt.get("statement_type"))
            units = stmt.get("units") or {}
            scale = units.get("scale")
            if scale:
                title = f"{title} (Scale: {scale})"
            title_cell = ws.cell(row=current_row, column=1, value=title)
            title_cell.font = header_font
            title_cell.fill = navy_fill
            title_cell.alignment = center_align
            current_row += 1

            for row in stmt.get("rows", []):
                label = row.get("label", "")
                values = row.get("values") or {}
                normalized_values = {}
                for key, val in values.items():
                    norm = normalize_year_key(key)
                    if norm:
                        normalized_values[norm] = val

                if row.get("is_section") or all(v is None for v in normalized_values.values()):
                    write_section_row(label)
                    continue

                val_row = [label]
                for year in sorted_years:
                    val_row.append(normalized_values.get(year))

                for col_idx, val in enumerate(val_row, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    if col_idx > 1 and isinstance(val, (int, float)):
                        cell.number_format = "#,##0.00"
                        cell.font = number_font
                        cell.alignment = center_align
                current_row += 1

            current_row += 2

    for col in ws.columns:
        max_len = 0
        numeric_col = False
        first_cell = col[0]
        col_idx = getattr(first_cell, "column", None)
        if col_idx is None:
            continue

        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
                if isinstance(cell.value, (int, float)):
                    numeric_col = True

        col_letter = get_column_letter(col_idx)
        width = max_len + (4 if numeric_col else 2)
        if numeric_col:
            width = max(width, 12)
        ws.column_dimensions[col_letter].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def create_excel_file(data_dict: Dict[str, Any]) -> BytesIO:
    """Wrapper to export the current statements dict to Excel."""
    statements = [stmt for stmt in data_dict.values() if stmt]
    return statements_to_excel(statements)


def main():
    # Header
    st.markdown('<div class="main-header">📊 Financial Statement Extractor</div>', unsafe_allow_html=True)
    st.markdown("---")
    
    # Sidebar for API keys and upload
    with st.sidebar:
        st.header("⚙️ Configuration")
        
        llama_api_key = st.text_input(
            "Parser API Key",
            type="password",
            help="Enter the API key for your PDF parser service"
        )
        
        openai_api_key = st.text_input(
            "Validator API Key",
            type="password",
            help="Enter the API key for your validator service"
        )
        
        st.markdown("---")
        
        st.header("📤 Upload PDF")
        uploaded_file = st.file_uploader(
            "Choose a PDF file",
            type=['pdf'],
            help="Upload a PDF containing financial statements"
        )
        
        if uploaded_file is not None:
            st.session_state.uploaded_file = uploaded_file
    
    # Main content area
    if not llama_api_key or not openai_api_key:
        st.info("👈 Please enter your API keys in the sidebar to begin.")
        return
    
    if st.session_state.uploaded_file is None:
        st.info("👈 Please upload a PDF file in the sidebar to begin extraction.")
        return
    
    st.markdown("### Live Processing Log")
    progress_placeholder = st.empty()
    progress_bar = progress_placeholder.progress(st.session_state.latest_progress)
    log_placeholder = st.empty()
    if st.session_state.latest_logs:
        log_placeholder.markdown("  \n".join(st.session_state.latest_logs))
    st.markdown("---")
    
    # Process button
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        process_button = st.button(
            "🚀 Extract Financial Statements",
            type="primary",
            use_container_width=True
        )
    
    # Process the PDF
    if process_button or st.session_state.processing_complete:
        if not st.session_state.processing_complete:
            st.session_state.latest_logs = []
            st.session_state.latest_progress = 0
            progress_bar.progress(0)
            log_placeholder.empty()

            status_icon_map = {
                "info": "ℹ️",
                "running": "⏳",
                "success": "✅",
                "warning": "⚠️",
                "error": "❌"
            }

            def ui_log(message, status="info"):
                icon = status_icon_map.get(status, "ℹ️")
                st.session_state.latest_logs.append(f"{icon} {message}")
                log_placeholder.markdown("  \n".join(st.session_state.latest_logs))

            def ui_progress(current_step, total_steps):
                total = max(total_steps, 1)
                percentage = int(min(max((current_step / total) * 100, 0), 100))
                st.session_state.latest_progress = percentage
                progress_bar.progress(percentage)

            # Save uploaded file temporarily
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
                tmp_file.write(st.session_state.uploaded_file.read())
                tmp_path = tmp_file.name
            
            try:
                # Run pipeline
                organized_data, error = process_pdf_pipeline(
                    tmp_path,
                    llama_api_key,
                    openai_api_key,
                    log_callback=ui_log,
                    progress_callback=ui_progress
                )
                
                if error:
                    st.error(f"❌ Error processing PDF: {error}")
                else:
                    st.session_state.extracted_data = organized_data
                    st.session_state.processing_complete = True
                    st.success("✅ Financial statements extracted successfully!")
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
            finally:
                # Clean up temp file
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)
        
        # Display results
        if st.session_state.processing_complete and st.session_state.extracted_data:
            data = st.session_state.extracted_data

            total_statements = sum(1 for v in data.values() if v)
            latest_period = "—"
            for stmt in data.values():
                if stmt:
                    periods = build_period_columns(stmt.get("periods"), stmt.get("rows"))
                    if periods:
                        latest_period = periods[-1]
                        break

            hero_html = f"""
            <div class="hero-card">
                <h2 style="margin-bottom:0.3rem;">Financial Data Ready</h2>
                <p style="margin-top:0;opacity:0.85;">Latest reporting period: <strong>{latest_period}</strong> • Statements captured: <strong>{total_statements}</strong></p>
            </div>
            """
            st.markdown(hero_html, unsafe_allow_html=True)

            metric_cols = st.columns(3)
            income_value = format_currency(extract_metric_value(data.get("income_statement"), ["revenue", "net sales", "total revenue"]))
            profit_value = format_currency(extract_metric_value(data.get("income_statement"), ["net income", "profit"]))
            cash_value = format_currency(extract_metric_value(data.get("cash_flow"), ["cash", "operating activities"]))

            with metric_cols[0]:
                st.markdown(f'<div class="metric-card"><div class="metric-label">Revenue (latest)</div><div class="metric-value">{income_value}</div></div>', unsafe_allow_html=True)
            with metric_cols[1]:
                st.markdown(f'<div class="metric-card"><div class="metric-label">Net Income (latest)</div><div class="metric-value">{profit_value}</div></div>', unsafe_allow_html=True)
            with metric_cols[2]:
                st.markdown(f'<div class="metric-card"><div class="metric-label">Cash Ops (latest)</div><div class="metric-value">{cash_value}</div></div>', unsafe_allow_html=True)
            
            # Download button
            st.markdown("---")
            col1, col2, col3 = st.columns([1, 1, 1])
            with col2:
                excel_file = create_excel_file(data)
                st.download_button(
                    label="📥 Download as Excel",
                    data=excel_file,
                    file_name="financial_statements.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            st.markdown("---")
            
            # Display Income Statement
            if data.get('income_statement'):
                st.markdown('<div class="statement-header">💰 Income Statement</div>', unsafe_allow_html=True)
                with st.expander("Show trend chart", expanded=True):
                    render_statement_chart(data['income_statement'], "Income Statement Trend")
                df_income = json_to_dataframe(data['income_statement'])
                if df_income is not None:
                    st.dataframe(
                        df_income,
                        use_container_width=True,
                        hide_index=True
                    )
                else:
                    st.warning("Income statement data is not in the expected format.")
            
            # Display Balance Sheet
            if data.get('balance_sheet'):
                st.markdown('<div class="statement-header">⚖️ Balance Sheet</div>', unsafe_allow_html=True)
                df_balance = json_to_dataframe(data['balance_sheet'])
                if df_balance is not None:
                    st.dataframe(
                        df_balance,
                        use_container_width=True,
                        hide_index=True
                    )
                else:
                    st.warning("Balance sheet data is not in the expected format.")
            
            # Display Cash Flow Statement
            if data.get('cash_flow'):
                st.markdown('<div class="statement-header">💸 Cash Flow Statement</div>', unsafe_allow_html=True)
                df_cashflow = json_to_dataframe(data['cash_flow'])
                if df_cashflow is not None:
                    st.dataframe(
                        df_cashflow,
                        use_container_width=True,
                        hide_index=True
                    )
                else:
                    st.warning("Cash flow statement data is not in the expected format.")
            
            # Show message if no data found
            if not any([data.get('income_statement'), data.get('balance_sheet'), data.get('cash_flow')]):
                st.warning("⚠️ No financial statements were found in the PDF. Please ensure the PDF contains income statements, balance sheets, or cash flow statements.")
            
            # Reset button
            st.markdown("---")
            if st.button("🔄 Process Another PDF", use_container_width=True):
                st.session_state.processing_complete = False
                st.session_state.extracted_data = {}
                st.session_state.uploaded_file = None
                st.rerun()

if __name__ == "__main__":
    main()

